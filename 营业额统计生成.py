#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
营业额统计表生成工具
======================
将「钉钉记录」和「日流水表」的数据合并，生成与统计表相同格式的 Excel 文件。

用法:
    python 营业额统计生成.py              # 启动 GUI 界面
    python 营业额统计生成.py --cli [...]  # 命令行模式

数据来源说明:
    - 收银记录(日流水)为权威数据源，原始财务流水
    - 钉钉记录为补充数据来源（店长手动上报）
    - 当两个表格同类型数据有差异(均有非零值但不一致)时，以收银记录(日流水)为准
    - 当收银记录某项为 0 而钉钉有值时，用钉钉补充（收银记录缺失而非差异）
    - 钉钉有但日流水没有的门店不写入统计表，仅在运行日志中列出
    - 如果钉钉记录有重复门店，用日流水数据判断哪条正确，自动去重
    - 营业额 = 各分类收入之和（现金+微信+支付宝+美团外卖+饿了么+京东外卖+美团团购+抖音团购+快手团购+其它外卖+其他收入）
"""

import sys
import os
import re
import argparse
from datetime import datetime, date

import xlrd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ============================================================
# 工具函数
# ============================================================

# 脚本/可执行文件所在目录（用于默认文件路径，让工具跟着程序走，不写死 D 盘）
# PyInstaller 打包后 __file__ 指向临时解压目录，需用 sys.executable（exe 所在路径）
if getattr(sys, 'frozen', False):
    SCRIPT_DIR = os.path.dirname(os.path.abspath(sys.executable))
else:
    SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))


def resource_path(filename):
    """获取开发环境或 PyInstaller 单文件环境中的资源路径。"""
    base_dir = getattr(sys, '_MEIPASS', SCRIPT_DIR)
    return os.path.join(base_dir, filename)


APP_ICON_PATH = resource_path('app-icon.ico')


def to_num(v):
    """将各种格式的值转换为 float，无效值返回 0。"""
    if v is None:
        return 0.0
    s = str(v).strip()
    if s == '' or s == '-' or s == 'nan' or s == 'None':
        return 0.0
    try:
        return float(s)
    except (ValueError, TypeError):
        return 0.0


def extract_store_id(text):
    """从门店名称中提取门店编号（如 YM010X, FJ011X, HY001 等）。"""
    if not text:
        return None
    m = re.search(r'[A-Z]{2,3}\d{3}[A-Z]?', str(text))
    return m.group() if m else None


def clean_store_name(name):
    """去除门店名称中的编号后缀、停业标记和多余空白。"""
    if not name:
        return ''
    s = str(name).strip()
    # 去除编号后缀: "中垌店-YM010X" -> "中垌店"
    s = re.sub(r'\s*[-]?\s*[A-Z]{2,3}\d{3}[A-Z]?\s*$', '', s)
    # 去除停业标记
    s = re.sub(r'[（(]停业[)）]', '', s)
    # 去除末尾句点
    s = re.sub(r'\.$', '', s)
    return s.strip()


def num_or_none(v):
    """数值为 0 时返回 None（与原统计表风格一致，空值不显示）。"""
    if v == 0 or v is None:
        return None
    return round(float(v), 2)


# 统计表列名
TONGJI_COLS = [
    '营业额', '现金', '微信', '支付宝',
    '美团外卖', '饿了么', '京东外卖',
    '美团团购', '抖音团购', '快手团购',
    '其它外卖', '其他收入',
]

# 低营业额门店标黄阈值（不含该金额本身）
LOW_REVENUE_THRESHOLD = 2000

# 金额严格大于以下阈值时判定为异常；等于阈值不提醒。
ABNORMAL_THRESHOLDS = {
    '营业额': 100000,
    '美团外卖': 20000,
    '饿了么': 20000,
    '京东外卖': 5000,
    '美团团购': 5000,
    '抖音团购': 20000,
    '快手团购': 5000,
    '其它外卖': 3000,
    '其他收入': 2000,
}

ABNORMAL_DISPLAY_NAMES = {
    '京东外卖': '京东',
    '快手团购': '快手',
    '其他收入': '其它收入',
}


def calc_total(store):
    """营业额 = 各分类收入之和。"""
    return sum(store.get(c, 0) for c in TONGJI_COLS if c != '营业额')


def get_abnormal_items(store):
    """返回门店中严格超过异常阈值的项目。"""
    items = []
    for name, threshold in ABNORMAL_THRESHOLDS.items():
        value = to_num(store.get(name, 0))
        if value > threshold:
            items.append({
                'name': name,
                'label': ABNORMAL_DISPLAY_NAMES.get(name, name),
                'value': value,
                'threshold': threshold,
            })
    return items


# ============================================================
# 读取钉钉记录
# ============================================================

def read_dingding(filepath):
    """
    读取钉钉记录.xlsx，提取每家门店的营业数据。

    钉钉记录结构:
    - 每条记录约8行（1行主数据 + 7行班次明细）
    - 但有少量单行记录（IT部门等无门店数据的记录），会打乱8行节奏
    - 解决方案: 扫描所有行，找到 col 1（提交人）非空的行作为主数据行
    """
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb.active

    stores = []

    # 扫描所有行，找到主数据行（col 1 即"提交人"非空）
    for r in range(3, ws.max_row + 1):
        submitter = ws.cell(row=r, column=1).value
        if submitter is None or str(submitter).strip() == '':
            continue  # 非主数据行（班次明细等）

        row = [ws.cell(row=r, column=c).value for c in range(1, ws.max_column + 1)]

        # 选择区 (col 9, 0-indexed=8)
        region = str(row[8]).strip() if row[8] else ''

        # 门店名称在 cols 10-19 (0-indexed 9-18) 中找非 '-' 的值
        store_field = None
        for ci in range(9, 19):
            val = row[ci]
            if val and str(val).strip() not in ('-', '', 'nan', 'None'):
                store_field = str(val).strip()
                break

        if not store_field:
            continue  # 无门店信息（IT部门等），跳过

        # 分离编号和门店名: "FJ011X 云霄2店" -> id=FJ011X, name=云霄2店
        store_id = extract_store_id(store_field)
        parts = store_field.split(None, 1)  # 按空白分割，最多2段
        if len(parts) == 2:
            store_name = parts[1].strip()
        else:
            store_name = store_field

        # 提取日期 (col 20, 0-indexed=19)
        date_str = str(row[19]).strip() if row[19] else ''

        # 各分类营业额
        def g(i):
            return to_num(row[i]) if i < len(row) else 0.0

        # 钉钉列映射 (0-indexed):
        # 总营业额=20, 现金=21, 微信=22, 支付宝=23
        # 美团外卖=24, 饿了么=25, 京东外卖=26
        # 美团团购=27, 抖音团购=28, 快手团购=29
        # 本地外卖1金额=32, 本地外卖2金额=35, 本地外卖3金额=38
        # 其他团购1金额=41, 其他团购2金额=44
        # 其他营业收入1金额=47, 其他营业收入2金额=50
        other_wm = g(32) + g(35) + g(38)
        other_income = g(41) + g(44) + g(47) + g(50)

        store = {
            'region': region,
            'store_name': store_name,
            'store_id': store_id,
            'source': '钉钉',
            'date_str': date_str,
            '现金': g(21),
            '微信': g(22),
            '支付宝': g(23),
            '美团外卖': g(24),
            '饿了么': g(25),
            '京东外卖': g(26),
            '美团团购': g(27),
            '抖音团购': g(28),
            '快手团购': g(29),
            '其它外卖': other_wm,
            '其他收入': other_income,
        }
        store['营业额'] = calc_total(store)
        stores.append(store)

    wb.close()
    return stores


# ============================================================
# 读取日流水表
# ============================================================

def _load_sheet(filepath):
    """
    读取 Excel 文件(.xls 或 .xlsx)，返回统一的二维数据列表。

    返回: list[list[值]], 跳过表头(第1行)
    """
    ext = os.path.splitext(filepath)[1].lower()
    if ext == '.xls':
        wb = xlrd.open_workbook(filepath, ignore_workbook_corruption=True)
        ws = wb.sheet_by_index(0)
        rows = []
        for r in range(ws.nrows):
            rows.append([ws.cell_value(r, c) for c in range(ws.ncols)])
        wb = None
    else:
        # .xlsx (openpyxl 读取所有数据行)
        wb = openpyxl.load_workbook(filepath, data_only=True)
        ws = wb.active
        rows = []
        for r in range(1, ws.max_row + 1):
            rows.append([ws.cell(row=r, column=c).value
                         for c in range(1, ws.max_column + 1)])
        wb.close()
    return rows


def read_riliushui(filepath):
    """
    读取日流水文件(.xls 或 .xlsx)，提取门店营业数据。

    日流水列映射 (0-indexed):
    - 区域=1, 门店=3, 应收金额=4
    - 现金=18, 微信=21, 支付宝=23
    - 抖音团购=26, 快手团购=29
    - 本地外卖=36, 美团团购r=37
    - 美团(平台收入)=51, 饿了么(平台收入)=49
    - 京东=55
    """
    rows = _load_sheet(filepath)

    stores = []
    for r in range(1, len(rows)):
        row = rows[r]
        if len(row) < 5:
            continue
        area = row[1]                       # 区域
        store_raw = row[3]                  # 门店

        if not store_raw or str(store_raw).strip() == '':
            continue

        # 跳过停业门店和 Test 门店
        area_str = str(area).strip() if area else ''
        if area_str in ('停业门店', 'Test'):
            continue

        # 跳过营业额为 0 的门店
        total_rev = to_num(row[4] if len(row) > 4 else None)
        if total_rev == 0:
            continue

        store_id = extract_store_id(store_raw)
        store_name = clean_store_name(store_raw)

        def g(c):
            return to_num(row[c] if c < len(row) else None)

        store = {
            'region': area_str,
            'store_name': store_name,
            'store_id': store_id,
            'source': '日流水',
            'date_str': '',
            '现金': g(18),                    # 现金
            '微信': g(21),                    # 微信
            '支付宝': g(23),                  # 支付宝
            '美团外卖': g(51),               # 美团(平台收入)
            '饿了么': g(49),                 # 饿了么(平台收入)
            '京东外卖': g(55),               # 京东
            '美团团购': g(37),               # 美团团购r
            '抖音团购': g(26),               # 抖音团购
            '快手团购': g(29),               # 快手团购
            '其它外卖': g(36),               # 本地外卖
            '其他收入': 0.0,
        }
        # 营业额 = 各分类收入之和
        store['营业额'] = calc_total(store)
        stores.append(store)

    return stores


# ============================================================
# 合并数据
# ============================================================

def normalize_id(sid):
    """归一化门店编号：去除尾部 X（YM013X → YM013），用于跨表匹配。"""
    if not sid:
        return None
    return sid.rstrip('X')


def normalize_name(name):
    """归一化门店名称：去除尾部"店"字，用于跨表匹配（"广海店" → "广海"）。"""
    if not name:
        return ''
    s = str(name).strip()
    if s.endswith('店'):
        s = s[:-1]
    return s


def merge_stores(dd_stores, rl_stores):
    """
    合并钉钉和日流水(收银记录)数据。

    优先级:
    1. 收银记录(日流水)是权威数据源，原始财务流水
    2. 当两个表格同类型数据有差异(均有非零值但不一致)时，以收银记录为准
    3. 当收银记录某项为 0 而钉钉有值时，用钉钉补充（收银记录缺失而非差异）
    4. 钉钉有但日流水没有的门店不写入统计表，仅返回供日志展示
    5. 如果钉钉有重复门店(同编号或同名)，用日流水数据判断哪条记录正确

    匹配规则:
    - 优先用归一化门店编号匹配(去尾部X): MM004X ≈ MM004
    - 门店名称兜底(去尾部"店"): "广海店" ≈ "广海"

    Returns:
        (merged, extra_dd) 元组
        - merged: 日流水门店（含钉钉补零），写入统计表
        - extra_dd: 钉钉有但日流水没有的门店，仅供日志展示
    """
    # 建立日流水索引
    rl_by_nid = {}     # 归一化编号 -> store
    rl_by_name = {}    # 门店名 -> store
    rl_by_nname = {}   # 归一化名 -> store

    for s in rl_stores:
        nid = normalize_id(s['store_id']) if s['store_id'] else None
        if nid:
            rl_by_nid[nid] = s
        if s['store_name']:
            rl_by_name[s['store_name']] = s
            rl_by_nname[normalize_name(s['store_name'])] = s

    # ---- 第一步: 钉钉去重 ----
    dd_deduped = _dedup_dingding(dd_stores, rl_by_nid, rl_by_name, rl_by_nname)

    # ---- 第二步: 以日流水为主，钉钉补零 ----
    merged = []
    for rl_s in rl_stores:
        merged.append(dict(rl_s))

    # 从 merged 副本重建查找字典，确保 fill-0 修改的是输出中的副本而非原件
    rl_by_nid.clear()
    rl_by_name.clear()
    rl_by_nname.clear()
    for s in merged:
        nid = normalize_id(s['store_id']) if s['store_id'] else None
        if nid:
            rl_by_nid[nid] = s
        if s['store_name']:
            rl_by_name[s['store_name']] = s
            rl_by_nname[normalize_name(s['store_name'])] = s

    # 钉钉补零: 已匹配的用钉钉填补日流水的 0 值
    # 未匹配的放入 extra_dd（不写入统计表）
    extra_dd = []
    for dd_s in dd_deduped:
        nid = normalize_id(dd_s['store_id']) if dd_s['store_id'] else None

        matched = None
        if nid and nid in rl_by_nid:
            matched = rl_by_nid[nid]
        elif dd_s['store_name'] in rl_by_name:
            matched = rl_by_name[dd_s['store_name']]
        elif normalize_name(dd_s['store_name']) in rl_by_nname:
            matched = rl_by_nname[normalize_name(dd_s['store_name'])]

        if matched:
            # 已在日流水中:
            # - 日流水有非零值 → 保留日流水值（以收银记录为准）
            # - 日流水为 0 而钉钉有值 → 用钉钉补充（收银记录缺失）
            for col in TONGJI_COLS:
                if col == '营业额':
                    continue
                if matched.get(col, 0) == 0 and dd_s.get(col, 0) != 0:
                    matched[col] = dd_s[col]
            # 重新计算营业额
            matched['营业额'] = calc_total(matched)
        else:
            # 日流水未覆盖，不写入统计表，记录到 extra_dd
            extra_dd.append(dict(dd_s))

    return merged, extra_dd


def _dedup_dingding(dd_stores, rl_by_nid, rl_by_name, rl_by_nname):
    """
    钉钉记录去重: 如果有重复门店(同编号或同名)，用日流水数据判断哪条正确。

    判断逻辑:
    - 找到日流水中对应的门店
    - 选择与日流水营业额差异最小的那条钉钉记录
    - 如果日流水无匹配，保留营业额最高的那条

    Returns:
        去重后的钉钉门店列表
    """
    # 按归一化编号分组
    groups = {}  # key -> [(index, store)]
    for i, s in enumerate(dd_stores):
        nid = normalize_id(s['store_id']) if s['store_id'] else None
        key = nid if nid else f'name:{s["store_name"]}'
        groups.setdefault(key, []).append((i, s))

    result = []
    for key, items in groups.items():
        if len(items) == 1:
            # 无重复，直接保留
            result.append(items[0][1])
        else:
            # 有重复，用日流水判断哪条正确
            nid = key if not key.startswith('name:') else None
            rl_match = None
            if nid and nid in rl_by_nid:
                rl_match = rl_by_nid[nid]
            else:
                for _, s in items:
                    if s['store_name'] in rl_by_name:
                        rl_match = rl_by_name[s['store_name']]
                        break
                    if normalize_name(s['store_name']) in rl_by_nname:
                        rl_match = rl_by_nname[normalize_name(s['store_name'])]
                        break

            if rl_match:
                # 选择与日流水营业额差异最小的那条
                best = min(items, key=lambda item: abs(item[1]['营业额'] - rl_match['营业额']))
                result.append(best[1])
            else:
                # 无日流水参考，保留营业额最高的
                best = max(items, key=lambda item: item[1]['营业额'])
                result.append(best[1])

    return result


# ============================================================
# 生成统计表
# ============================================================

def generate_tongji(stores, output_path, date_str):
    """
    生成与原统计表格式相同的 Excel 文件。

    格式特点:
    - 第1行: 标题（合并单元格，黄色背景，大字号粗体）
    - 第2行: 表头（黄色背景，加粗，居中）
    - 第3行起: 数据，按区域分组，合并区域列单元格
    - 各区域内按营业额降序排列
    - 营业额低于 2000 元的门店整行标黄（区域列除外）
    - 超过异常阈值的具体金额单元格标红
    - 合计行（黄色背景，SUM 公式，加粗）
    - 备注行（灰色斜体）
    """
    # 区域排序
    region_order = [
        '福建区', '河源区', '雷州区', '梅州区', '茂名区',
        '南油区', '阳江区', '阳茂区', '湛江区', '肇庆区',
    ]

    # 按区域分组
    region_groups = {}
    for s in stores:
        r = s['region']
        if r not in region_groups:
            region_groups[r] = []
        region_groups[r].append(s)

    # 区域内按营业额降序
    for r in region_groups:
        region_groups[r].sort(key=lambda x: x['营业额'], reverse=True)

    # 区域排序
    sorted_regions = [r for r in region_order if r in region_groups]
    for r in region_groups:
        if r not in sorted_regions:
            sorted_regions.append(r)

    # 创建工作簿
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '营业额统计'

    # ========== 样式定义（完全对照原表） ==========
    # 注意: 必须用 8 位 ARGB 格式, 6 位会被 openpyxl 补 00 导致透明
    YELLOW     = 'FFFFFF00'   # 纯黄 #FFFF00 (FF=不透明)
    ALERT_RED  = 'FFE53935'   # 异常红 #E53935
    BLACK      = 'FF000000'   # 黑色（用于字体）
    WHITE      = 'FFFFFFFF'   # 纯白（用于边框）

    yellow_fill = PatternFill(start_color=YELLOW, end_color=YELLOW, fill_type='solid')
    alert_fill = PatternFill(start_color=ALERT_RED, end_color=ALERT_RED, fill_type='solid')

    # 区域交替填充（对照参考表）: 浅橙 #FCE4D6 / 浅绿 #E2F0D9
    fill_a = PatternFill(fill_type='solid',
                         start_color='FFFCE4D6',
                         end_color='FFFCE4D6')
    fill_b = PatternFill(fill_type='solid',
                         start_color='FFE2F0D9',
                         end_color='FFE2F0D9')

    # 字体: 全部微软雅黑 12pt 黑色, 标题 16pt
    title_font  = Font(name='微软雅黑', bold=True,  size=16, color=BLACK)
    font_12b    = Font(name='微软雅黑', bold=True,  size=12, color=BLACK)  # 粗体
    font_12     = Font(name='微软雅黑', bold=False, size=12, color=BLACK)  # 常规
    font_12_alert = Font(name='微软雅黑', bold=True, size=12, color=WHITE)

    # 对齐
    center = Alignment(horizontal='center', vertical='center')
    left   = Alignment(horizontal='left',   vertical='center')

    # 边框: 全部 medium 纯白 (FFFFFFFF)
    side_med = Side(style='medium', color=WHITE)
    border_all = Border(left=side_med, right=side_med, top=side_med, bottom=side_med)

    # 表头 (14列)
    headers = [
        '选择区', '门店', '营业额', '现金', '微信', '支付宝',
        '美团外卖', '饿了么', '京东外卖', '美团团购',
        '抖音团购', '快手团购', '其它外卖', '其他收入',
    ]
    N_COLS = len(headers)
    LAST_COL = get_column_letter(N_COLS)   # 'N'

    # ========== 第1行: 标题 ==========
    # 先设样式再合并（合并后非首单元格变为只读 MergedCell）
    for col_idx in range(1, N_COLS + 1):
        cc = ws.cell(row=1, column=col_idx)
        cc.fill = yellow_fill
        cc.font = title_font
        cc.alignment = center
        cc.border = border_all
    ws.cell(row=1, column=1, value=f'{date_str}营业额统计')
    ws.merge_cells(f'A1:{LAST_COL}1')
    ws.row_dimensions[1].height = 30

    # ========== 第2行: 表头 ==========
    for col_idx, h in enumerate(headers, 1):
        c = ws.cell(row=2, column=col_idx, value=h)
        c.fill = yellow_fill
        c.font = font_12b
        c.alignment = center
        c.border = border_all
    ws.row_dimensions[2].height = 25

    # ========== 数据行 ==========
    cur_row = 3
    data_start = cur_row

    for region_idx, region in enumerate(sorted_regions):
        stores_in_region = region_groups[region]
        region_start = cur_row
        region_end = cur_row + len(stores_in_region) - 1
        # 交替填充: 奇数区域用 fill_a, 偶数区域用 fill_b
        region_fill = fill_a if region_idx % 2 == 0 else fill_b

        for i, store in enumerate(stores_in_region):
            is_low_revenue = store['营业额'] < LOW_REVENUE_THRESHOLD

            # 选择区(合并)
            ws.cell(row=cur_row, column=1, value=region)
            # 门店
            ws.cell(row=cur_row, column=2, value=store['store_name'])
            # 数值列(从第3列营业额开始)
            for ci, col_name in enumerate(TONGJI_COLS):
                val = num_or_none(store[col_name])
                ws.cell(row=cur_row, column=ci + 3, value=val)

            # 应用样式: 边框 + 字体 + 对齐 + 区域交替填充
            for col_idx in range(1, N_COLS + 1):
                cell = ws.cell(row=cur_row, column=col_idx)
                cell.border = border_all
                column_name = headers[col_idx - 1]
                is_abnormal = (
                    column_name in ABNORMAL_THRESHOLDS
                    and to_num(store.get(column_name, 0)) > ABNORMAL_THRESHOLDS[column_name]
                )
                # 低营业额门店从“门店”列到最后一个金额列整行标黄。
                # 区域列需要纵向合并，继续使用区域底色，避免合并后样式混乱。
                if is_abnormal:
                    cell.fill = alert_fill
                elif is_low_revenue and col_idx >= 2:
                    cell.fill = yellow_fill
                else:
                    cell.fill = region_fill
                if is_abnormal:
                    cell.font = font_12_alert
                    cell.alignment = center
                elif col_idx == 1:
                    cell.font = font_12b   # 区域列加粗
                    cell.alignment = center
                elif col_idx == 2:
                    cell.font = font_12    # 门店列常规
                    cell.alignment = left
                else:
                    cell.font = font_12    # 数值列常规
                    cell.alignment = center

            ws.row_dimensions[cur_row].height = 18
            cur_row += 1

        # 合并区域单元格
        if region_end > region_start:
            ws.merge_cells(f'A{region_start}:A{region_end}')

    data_end = cur_row - 1

    # ========== 合计行 ==========
    total_row = cur_row
    ws.cell(row=total_row, column=1, value=None)
    ws.cell(row=total_row, column=2, value='合计')
    # 数值列: 营业额/各分类用 SUM 公式
    for ci in range(3, N_COLS + 1):
        col_letter = get_column_letter(ci)
        ws.cell(row=total_row, column=ci,
                value=f'=SUM({col_letter}{data_start}:{col_letter}{data_end})')

    for col_idx in range(1, N_COLS + 1):
        c = ws.cell(row=total_row, column=col_idx)
        c.fill = yellow_fill
        c.font = font_12b
        c.border = border_all
        if col_idx == 2:
            c.alignment = left
        else:
            c.alignment = center
    ws.row_dimensions[total_row].height = 25
    cur_row += 1

    # ========== 备注行（黄底 12pt 粗体 居中） ==========
    # 先设样式再合并
    for col_idx in range(1, N_COLS + 1):
        cc = ws.cell(row=cur_row, column=col_idx)
        cc.fill = yellow_fill
        cc.font = font_12b
        cc.alignment = center
        cc.border = border_all
    ws.cell(row=cur_row, column=1,
            value='标黄门店表示统计当日,营业额低于2000餐厅,便于关注！')
    ws.merge_cells(f'A{cur_row}:{LAST_COL}{cur_row}')
    ws.row_dimensions[cur_row].height = 25
    cur_row += 1

    for col_idx in range(1, N_COLS + 1):
        cc = ws.cell(row=cur_row, column=col_idx)
        cc.fill = yellow_fill
        cc.font = font_12b
        cc.alignment = center
        cc.border = border_all
    ws.cell(row=cur_row, column=1,
            value='标红单元格表示该项金额超过异常阈值,请及时核对！')
    ws.merge_cells(f'A{cur_row}:{LAST_COL}{cur_row}')
    ws.row_dimensions[cur_row].height = 25
    cur_row += 1

    for col_idx in range(1, N_COLS + 1):
        cc = ws.cell(row=cur_row, column=col_idx)
        cc.fill = yellow_fill
        cc.font = font_12b
        cc.alignment = center
        cc.border = border_all
    ws.cell(row=cur_row, column=1,
            value='制表数据来自收银记录与店长钉钉上报,仅供参考,实收数据请以财务报表为准！')
    ws.merge_cells(f'A{cur_row}:{LAST_COL}{cur_row}')
    ws.row_dimensions[cur_row].height = 25

    # ========== 列宽（对照原表） ==========
    widths = {
        'A':  7.625,  # 选择区
        'B': 26.125,  # 门店
        'C': 10.375,  # 营业额
        'D':  9.25,   # 现金
        'E': 10.375,  # 微信
        'F':  8.125,  # 支付宝
        'G':  9.625,  # 美团外卖
        'H':  9.25,   # 饿了么
        'I':  9.625,  # 京东外卖
        'J': 13.0,    # 美团团购
        'K': 13.0,    # 抖音团购
        'L': 13.0,    # 快手团购
        'M': 10.375,  # 其它外卖
        'N':  9.625,  # 其他收入
    }
    for col_letter, w in widths.items():
        ws.column_dimensions[col_letter].width = w

    # 不冻结窗格（对照原表）

    wb.save(output_path)
    return output_path


# ============================================================
# 核心处理流程（CLI 与 GUI 共用）
# ============================================================

def run_pipeline(dingding_path, riliushui_path, output_path=None, date_str=None, log=print):
    """
    执行完整的读取→合并→生成流程。

    Args:
        dingding_path:  钉钉记录文件路径
        riliushui_path: 日流水文件路径
        output_path:    输出文件路径（None 时自动生成）
        date_str:       日期显示文字（None 时自动提取）
        log:            日志输出函数
    Returns:
        (output_path, store_count, region_count)
    """
    log(f'[1/4] 读取钉钉记录: {dingding_path}')
    if not os.path.exists(dingding_path):
        log(f'  错误: 文件不存在: {dingding_path}')
        return None, 0, 0
    dd_stores = read_dingding(dingding_path)
    log(f'  钉钉记录: {len(dd_stores)} 家门店')

    log(f'[2/4] 读取日流水: {riliushui_path}')
    if not os.path.exists(riliushui_path):
        log(f'  错误: 文件不存在: {riliushui_path}')
        return None, 0, 0
    rl_stores = read_riliushui(riliushui_path)
    log(f'  日流水: {len(rl_stores)} 家门店 (排除停业/零营业额)')

    log('[3/4] 合并数据...')
    # 先检测钉钉重复
    from collections import defaultdict
    dd_groups = defaultdict(list)
    for s in dd_stores:
        nid = normalize_id(s['store_id']) if s['store_id'] else f'name:{s["store_name"]}'
        dd_groups[nid].append(s)
    dup_count = sum(1 for v in dd_groups.values() if len(v) > 1)
    if dup_count:
        log(f'  钉钉发现 {dup_count} 组重复门店，已用日流水数据判断去重')

    merged, extra_dd = merge_stores(dd_stores, rl_stores)
    log(f'  以日流水为准: {len(merged)} 家门店')
    if extra_dd:
        log(f'  ┌─ 钉钉有但日流水没有的门店 ({len(extra_dd)} 家，仅供参考):')
        extra_by_region = {}
        for s in extra_dd:
            r = s.get('region', '未知')
            extra_by_region.setdefault(r, []).append(s)
        for r, stores in sorted(extra_by_region.items()):
            names = '、'.join(s['store_name'] for s in stores)
            log(f'  │  {r}: {names}')
        log(f'  └─ （以上门店未写入统计表）')

    abnormal_stores = []
    for store in merged:
        items = get_abnormal_items(store)
        if items:
            abnormal_stores.append((store, items))
    if abnormal_stores:
        abnormal_count = sum(len(items) for _, items in abnormal_stores)
        log(f'[异常提醒] 发现 {len(abnormal_stores)} 家门店、{abnormal_count} 项金额超过阈值')
        for store, items in abnormal_stores:
            details = '；'.join(
                f'{item["label"]} {item["value"]:,.2f}（阈值 {item["threshold"]:,.0f}）'
                for item in items
            )
            log(f'[异常提醒] {store["store_name"]}：{details}')
    else:
        log('  异常检查通过，未发现超阈值项目')

    region_counts = {}
    for s in merged:
        r = s['region']
        region_counts[r] = region_counts.get(r, 0) + 1
    for r, cnt in sorted(region_counts.items()):
        log(f'    {r}: {cnt} 家')

    # 确定日期
    if date_str:
        pass
    else:
        date_str = ''
        for s in dd_stores:
            if s.get('date_str'):
                date_str = s['date_str']
                break
        if date_str:
            try:
                dt = datetime.strptime(date_str, '%Y-%m-%d')
                date_str = f'{dt.month}月{dt.day}日'
            except ValueError:
                pass
        else:
            today = date.today()
            date_str = f'{today.month}月{today.day}日'

    # 确定输出路径
    if not output_path:
        for s in dd_stores:
            if s.get('date_str'):
                raw_date = s['date_str']
                break
        else:
            raw_date = date_str
        output_path = os.path.join(SCRIPT_DIR, f'{raw_date} 统计表.xlsx')

    log(f'[4/4] 生成统计表: {output_path}')
    generate_tongji(merged, output_path, date_str)
    log(f'\n完成！输出文件: {output_path}')
    log(f'共 {len(merged)} 家门店，{len(region_counts)} 个区域')

    return output_path, len(merged), len(region_counts)


# ============================================================
# CLI 模式
# ============================================================

def cli_main():
    parser = argparse.ArgumentParser(
        description='营业额统计表生成工具 — 合并钉钉记录和日流水数据生成统计表',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
示例:
  python 营业额统计生成.py --cli
  python 营业额统计生成.py --cli --date 8月3日
  python 营业额统计生成.py --cli --dingding 钉钉.xlsx --riliushui 日流水.xls --output 输出.xlsx
        """,
    )
    parser.add_argument('--cli', action='store_true',
                        help='使用命令行模式（不加此参数则启动 GUI）')
    parser.add_argument('--dingding', default=os.path.join(SCRIPT_DIR, '钉钉记录.xlsx'),
                        help='钉钉记录文件路径')
    parser.add_argument('--riliushui', default=os.path.join(SCRIPT_DIR, '日流水.xls'),
                        help='日流水文件路径')
    parser.add_argument('--output', default=None,
                        help='输出文件路径')
    parser.add_argument('--date', default=None,
                        help='统计日期显示文字，如 "8月2日"')

    args = parser.parse_args()

    output_path, n_stores, n_regions = run_pipeline(
        args.dingding, args.riliushui, args.output, args.date
    )
    if output_path is None:
        sys.exit(1)


# ============================================================
# GUI 模式 (tkinter)
# ============================================================

def gui_main():
    import tkinter as tk
    from tkinter import ttk, filedialog, messagebox
    import threading
    import queue

    root = tk.Tk()
    root.title('营业额统计表生成工具')
    root.resizable(False, False)

    # 居中显示
    window_width = 760
    window_height = 680
    screen_w = root.winfo_screenwidth()
    screen_h = root.winfo_screenheight()
    x = (screen_w - window_width) // 2
    y = (screen_h - window_height) // 2
    root.geometry(f'{window_width}x{window_height}+{x}+{y}')

    # --- 变量 ---
    dd_var = tk.StringVar()
    rl_var = tk.StringVar()
    date_var = tk.StringVar()
    output_var = tk.StringVar()
    status_var = tk.StringVar(value='准备就绪')
    generating = {'state': False}
    log_queue = queue.Queue()

    # 默认值：指向软件所在目录（不再写死 D 盘）
    dd_var.set(os.path.join(SCRIPT_DIR, '钉钉记录.xlsx'))
    rl_var.set(os.path.join(SCRIPT_DIR, '日流水.xls'))

    # --- QQ 风格浅蓝/淡紫渐变背景 ---
    BG_TOP = '#DDF4FF'
    BG_BOTTOM = '#F9EEFF'
    CARD_BG = '#FBFDFF'
    TEXT = '#172033'
    MUTED = '#7C879B'
    PRIMARY = '#4D8DFF'

    canvas = tk.Canvas(root, width=window_width, height=window_height,
                       highlightthickness=0, bd=0)
    canvas.pack(fill='both', expand=True)

    def hex_to_rgb(color):
        color = color.lstrip('#')
        return tuple(int(color[i:i + 2], 16) for i in (0, 2, 4))

    def mix_color(start, end, ratio):
        a = hex_to_rgb(start)
        b = hex_to_rgb(end)
        values = [round(a[i] + (b[i] - a[i]) * ratio) for i in range(3)]
        return '#{:02x}{:02x}{:02x}'.format(*values)

    # 用窄色带绘制渐变，不依赖 Pillow 等第三方图形库。
    for y in range(0, window_height, 2):
        ratio = y / max(window_height - 1, 1)
        canvas.create_rectangle(0, y, window_width, y + 2,
                                fill=mix_color(BG_TOP, BG_BOTTOM, ratio),
                                outline='')

    def rounded_rectangle(x1, y1, x2, y2, radius=24, **kwargs):
        points = [
            x1 + radius, y1, x2 - radius, y1,
            x2, y1, x2, y1 + radius,
            x2, y2 - radius, x2, y2,
            x2 - radius, y2, x1 + radius, y2,
            x1, y2, x1, y2 - radius,
            x1, y1 + radius, x1, y1,
        ]
        return canvas.create_polygon(points, smooth=True, splinesteps=24, **kwargs)

    # 顶部双环标识，呼应参考图的轻盈品牌感。
    canvas.create_oval(353, 23, 379, 49, outline='#57C7FF', width=3)
    canvas.create_oval(374, 23, 400, 49, outline='#A88DFF', width=3)
    canvas.create_text(window_width // 2, 73, text='营业额统计',
                       font=('Microsoft YaHei UI', 20, 'bold'), fill=TEXT)
    canvas.create_text(window_width // 2, 104,
                       text='合并钉钉记录与日流水，一键生成统计表',
                       font=('Microsoft YaHei UI', 10), fill=MUTED)

    # 卡片阴影与主体。
    rounded_rectangle(27, 124, 733, 658, radius=26,
                      fill='#D8E2F1', outline='')
    rounded_rectangle(24, 120, 730, 654, radius=26,
                      fill=CARD_BG, outline='#EAF0F8', width=1)
    card = tk.Frame(canvas, bg=CARD_BG, bd=0, highlightthickness=0)
    canvas.create_window(window_width // 2, 387, window=card,
                         width=672, height=506)

    # clam 主题可以在不同 Windows 版本上稳定呈现自定义配色。
    style = ttk.Style(root)
    try:
        style.theme_use('clam')
    except tk.TclError:
        pass
    style.configure('Card.TFrame', background=CARD_BG)
    style.configure('Field.TLabel', background=CARD_BG, foreground=TEXT,
                    font=('Microsoft YaHei UI', 10, 'bold'))
    style.configure('Hint.TLabel', background=CARD_BG, foreground=MUTED,
                    font=('Microsoft YaHei UI', 9))
    style.configure('Status.TLabel', background=CARD_BG, foreground=MUTED,
                    font=('Microsoft YaHei UI', 9))
    style.configure('Modern.TEntry', font=('Microsoft YaHei UI', 10),
                    padding=(10, 7), fieldbackground='#FFFFFF',
                    foreground=TEXT, bordercolor='#D9E4F2',
                    lightcolor='#D9E4F2', darkcolor='#D9E4F2')
    style.map('Modern.TEntry',
              bordercolor=[('focus', '#77A9FF')],
              lightcolor=[('focus', '#77A9FF')],
              darkcolor=[('focus', '#77A9FF')])
    style.configure('Browse.TButton', font=('Microsoft YaHei UI', 9),
                    foreground=PRIMARY, background='#EDF5FF',
                    borderwidth=0, padding=(13, 7), relief='flat')
    style.map('Browse.TButton',
              background=[('active', '#DCEBFF'), ('pressed', '#CFE3FF')])
    style.configure('Accent.TButton', font=('Microsoft YaHei UI', 11, 'bold'),
                    foreground='#FFFFFF', background=PRIMARY,
                    borderwidth=0, padding=(28, 9), relief='flat')
    style.map('Accent.TButton',
              background=[('active', '#397CF1'), ('pressed', '#2F6FDD'),
                          ('disabled', '#AFC9F5')],
              foreground=[('disabled', '#F6F9FF')])
    style.configure('Card.TLabelframe', background=CARD_BG,
                    bordercolor='#E3EAF4', lightcolor='#E3EAF4',
                    darkcolor='#E3EAF4', borderwidth=1, relief='solid')
    style.configure('Card.TLabelframe.Label', background=CARD_BG,
                    foreground=TEXT, font=('Microsoft YaHei UI', 9, 'bold'))

    # --- 钉钉记录 ---
    row1 = ttk.Frame(card, style='Card.TFrame', padding=(18, 18, 18, 5))
    row1.pack(fill='x')
    ttk.Label(row1, text='钉钉记录', width=10,
              style='Field.TLabel').pack(side='left')
    dd_entry = ttk.Entry(row1, textvariable=dd_var, style='Modern.TEntry')
    dd_entry.pack(side='left', fill='x', expand=True, padx=(4, 8))
    ttk.Button(row1, text='浏览', style='Browse.TButton',
               command=lambda: select_file(dd_var, '钉钉记录', SCRIPT_DIR)).pack(side='left')

    # --- 日流水表 ---
    row2 = ttk.Frame(card, style='Card.TFrame', padding=(18, 5, 18, 5))
    row2.pack(fill='x')
    ttk.Label(row2, text='日流水表', width=10,
              style='Field.TLabel').pack(side='left')
    rl_entry = ttk.Entry(row2, textvariable=rl_var, style='Modern.TEntry')
    rl_entry.pack(side='left', fill='x', expand=True, padx=(4, 8))
    ttk.Button(row2, text='浏览', style='Browse.TButton',
               command=lambda: select_file(rl_var, '日流水表', SCRIPT_DIR)).pack(side='left')

    # --- 日期行 ---
    row3 = ttk.Frame(card, style='Card.TFrame', padding=(18, 5, 18, 5))
    row3.pack(fill='x')
    ttk.Label(row3, text='统计日期', width=10,
              style='Field.TLabel').pack(side='left')
    ttk.Entry(row3, textvariable=date_var, width=23,
              style='Modern.TEntry').pack(side='left', padx=(4, 10))
    ttk.Label(row3, text='留空将从钉钉记录自动提取',
              style='Hint.TLabel').pack(side='left')

    # --- 输出路径行 ---
    row4 = ttk.Frame(card, style='Card.TFrame', padding=(18, 5, 18, 8))
    row4.pack(fill='x')
    ttk.Label(row4, text='输出路径', width=10,
              style='Field.TLabel').pack(side='left')
    out_entry = ttk.Entry(row4, textvariable=output_var, style='Modern.TEntry')
    out_entry.pack(side='left', fill='x', expand=True, padx=(4, 8))
    ttk.Button(row4, text='浏览', style='Browse.TButton',
               command=lambda: select_save_file(output_var, SCRIPT_DIR)).pack(side='left')

    # --- 生成按钮 ---
    btn_frame = ttk.Frame(card, style='Card.TFrame', padding=(18, 5, 18, 7))
    btn_frame.pack(fill='x')
    action_inner = ttk.Frame(btn_frame, style='Card.TFrame')
    action_inner.pack(anchor='center')
    gen_btn = ttk.Button(action_inner, text='生成统计表', style='Accent.TButton',
                         command=lambda: on_generate(root))
    gen_btn.pack(side='left')
    ttk.Label(action_inner, textvariable=status_var,
              style='Status.TLabel').pack(side='left', padx=(12, 0))

    # --- 日志区 ---
    log_frame = ttk.LabelFrame(card, text=' 运行日志 ', style='Card.TLabelframe',
                               padding=(10, 7))
    log_frame.pack(fill='both', expand=True, padx=20, pady=(4, 18))

    log_text = tk.Text(log_frame, font=('Microsoft YaHei UI', 9), height=10,
                       wrap='word', bg='#F5F8FC', fg='#455269',
                       insertbackground=TEXT, selectbackground='#CFE3FF',
                       relief='flat', bd=0, padx=10, pady=8,
                       highlightthickness=0)
    log_scroll = ttk.Scrollbar(log_frame, command=log_text.yview)
    log_text.configure(yscrollcommand=log_scroll.set)
    log_scroll.pack(side='right', fill='y')
    log_text.pack(side='left', fill='both', expand=True)
    log_text.tag_configure('error', foreground='#D94A5A')
    log_text.tag_configure('alert', foreground='#D82323',
                           font=('Microsoft YaHei UI', 9, 'bold'))
    log_text.tag_configure('success', foreground='#31845B')
    log_text.configure(state='disabled')

    def select_file(var, label, init_dir):
        path = filedialog.askopenfilename(
            title=f'选择{label}文件',
            initialdir=init_dir,
            filetypes=[
                ('Excel 文件', '*.xlsx *.xls'),
                ('所有文件', '*.*'),
            ],
        )
        if path:
            var.set(path)

    def select_save_file(var, init_dir):
        path = filedialog.asksaveasfilename(
            title='选择输出文件位置',
            initialdir=init_dir,
            defaultextension='.xlsx',
            filetypes=[('Excel 文件', '*.xlsx')],
        )
        if path:
            var.set(path)

    def append_log(msg):
        """仅在 GUI 主线程中更新日志控件。"""
        msg = str(msg)
        tag = ''
        if msg.startswith('[异常提醒]'):
            tag = 'alert'
        elif '错误' in msg or '失败' in msg:
            tag = 'error'
        elif '完成' in msg:
            tag = 'success'
        log_text.configure(state='normal')
        log_text.insert('end', msg + '\n', tag)
        log_text.see('end')
        log_text.configure(state='disabled')

    def log(msg):
        """后台线程只写队列，避免跨线程直接操作 Tkinter 控件。"""
        log_queue.put(msg)

    def poll_log_queue():
        try:
            while True:
                append_log(log_queue.get_nowait())
        except queue.Empty:
            pass
        root.after(80, poll_log_queue)

    def show_success(out, n, regions):
        output_var.set(out)
        status_var.set(f'完成 · {n} 家门店 / {regions} 个区域')
        messagebox.showinfo(
            '生成完成',
            f'统计表已生成：\n{out}\n\n共 {n} 家门店，{regions} 个区域',
        )

    def show_failure(message):
        status_var.set('生成失败，请查看运行日志')
        messagebox.showerror('生成失败', message)

    def finish_generation():
        gen_btn.configure(state='normal', text='生成统计表')
        generating['state'] = False

    def on_generate(root):
        if generating['state']:
            return
        dd_path = dd_var.get().strip()
        rl_path = rl_var.get().strip()
        out_path = output_var.get().strip() or None
        d_str = date_var.get().strip() or None

        if not dd_path:
            messagebox.showwarning('提示', '请先选择钉钉记录文件')
            return
        if not os.path.exists(dd_path):
            messagebox.showerror('错误', f'钉钉记录文件不存在：\n{dd_path}')
            return
        if not rl_path:
            messagebox.showwarning('提示', '请先选择日流水文件')
            return
        if not os.path.exists(rl_path):
            messagebox.showerror('错误', f'日流水文件不存在：\n{rl_path}')
            return

        # 清空日志
        log_text.configure(state='normal')
        log_text.delete('1.0', 'end')
        log_text.configure(state='disabled')

        gen_btn.configure(state='disabled', text='正在生成…')
        status_var.set('正在读取并合并数据…')
        generating['state'] = True

        def task():
            try:
                result = run_pipeline(dd_path, rl_path, out_path, d_str, log=log)
                out, n, regions = result
                # 用默认参数固化变量值,避免闭包作用域问题
                if out:
                    root.after(0, lambda out=out, n=n, regions=regions:
                        show_success(out, n, regions))
                else:
                    root.after(0, lambda: show_failure('生成失败，请查看运行日志'))
            except Exception as e:
                log(f'\n错误: {e}')
                # 注意: e 在 except 块退出后会被删除,lambda 必须用默认参数固化
                root.after(0, lambda err=e: show_failure(str(err)))
            finally:
                root.after(0, finish_generation)

        threading.Thread(target=task, daemon=True).start()

    root.after(80, poll_log_queue)
    root.mainloop()


def gui_main_fluent():
    """Win11 Fluent / Mica 风格界面，业务处理仍复用 run_pipeline。"""
    import tkinter as tk
    from tkinter import ttk, filedialog, messagebox
    import threading
    import queue
    import ctypes

    # 高 DPI 模式必须在创建 Tk 窗口之前启用。
    if sys.platform == 'win32':
        try:
            ctypes.windll.user32.SetProcessDpiAwarenessContext(ctypes.c_void_p(-4))
        except (AttributeError, OSError):
            try:
                ctypes.windll.shcore.SetProcessDpiAwareness(2)
            except (AttributeError, OSError):
                pass

    root = tk.Tk()
    root.title('营业额统计')
    root.resizable(False, False)

    window_width = 920
    window_height = 720
    screen_w = root.winfo_screenwidth()
    screen_h = root.winfo_screenheight()
    x = (screen_w - window_width) // 2
    y = (screen_h - window_height) // 2
    root.geometry(f'{window_width}x{window_height}+{x}+{y}')

    # 增加真实透明度；保持在 96.5%，避免文字和控件也变得发灰。
    try:
        root.attributes('-alpha', 0.965)
    except tk.TclError:
        pass

    def enable_windows_11_effects():
        """启用 Win11 原生圆角和 Mica，旧系统自动忽略。"""
        if sys.platform != 'win32':
            return
        try:
            root.update_idletasks()
            hwnd = ctypes.windll.user32.GetParent(root.winfo_id())
            corner_preference = ctypes.c_int(2)  # DWMWCP_ROUND
            backdrop_type = ctypes.c_int(2)      # DWMSBT_MAINWINDOW
            ctypes.windll.dwmapi.DwmSetWindowAttribute(
                hwnd, 33, ctypes.byref(corner_preference),
                ctypes.sizeof(corner_preference)
            )
            ctypes.windll.dwmapi.DwmSetWindowAttribute(
                hwnd, 38, ctypes.byref(backdrop_type),
                ctypes.sizeof(backdrop_type)
            )
        except (AttributeError, OSError):
            pass

    root.after(0, enable_windows_11_effects)

    dd_var = tk.StringVar(value=os.path.join(SCRIPT_DIR, '钉钉记录.xlsx'))
    rl_var = tk.StringVar(value=os.path.join(SCRIPT_DIR, '日流水.xls'))
    date_var = tk.StringVar()
    output_var = tk.StringVar()
    status_var = tk.StringVar(value='准备就绪')
    action_status_var = tk.StringVar(value='等待开始')
    generating = {'state': False}
    log_queue = queue.Queue()

    # Fluent 设计令牌。
    BG_TOP = '#E9F4FF'
    BG_BOTTOM = '#EEF1FF'
    CARD_BG = '#F2F8FD'
    FIELD_BG = '#FAFCFF'
    LOG_BG = '#F5F9FD'
    TEXT = '#172033'
    MUTED = '#667085'
    PRIMARY = '#2563EB'
    PRIMARY_HOVER = '#1D4ED8'
    BORDER = '#CFDAE8'
    FOCUS = '#3B82F6'
    SUCCESS = '#248A4B'

    canvas = tk.Canvas(root, width=window_width, height=window_height,
                       highlightthickness=0, bd=0, bg=BG_TOP)
    canvas.pack(fill='both', expand=True)

    def hex_to_rgb(color):
        color = color.lstrip('#')
        return tuple(int(color[i:i + 2], 16) for i in (0, 2, 4))

    def mix_color(start, end, ratio):
        a = hex_to_rgb(start)
        b = hex_to_rgb(end)
        values = [round(a[i] + (b[i] - a[i]) * ratio) for i in range(3)]
        return '#{:02x}{:02x}{:02x}'.format(*values)

    for line_y in range(0, window_height, 2):
        ratio = line_y / max(window_height - 1, 1)
        canvas.create_rectangle(
            0, line_y, window_width, line_y + 2,
            fill=mix_color(BG_TOP, BG_BOTTOM, ratio), outline=''
        )

    def radial_blob(cx, cy, radius, core_color, base_color, steps=38):
        """用同心色层模拟 Acrylic 背景后的柔焦光斑。"""
        for step in range(steps, 0, -1):
            ratio = step / steps
            radius_now = radius * ratio
            strength = (1.0 - ratio) * 0.50 + 0.035
            color = mix_color(base_color, core_color, strength)
            canvas.create_oval(
                cx - radius_now, cy - radius_now,
                cx + radius_now, cy + radius_now,
                fill=color, outline=''
            )

    radial_blob(82, 585, 270, '#C5B3FF', BG_BOTTOM)
    radial_blob(842, 615, 300, '#9DEAFF', BG_BOTTOM)
    radial_blob(650, 82, 240, '#C9E2FF', BG_TOP)

    def rounded_rectangle(x1, y1, x2, y2, radius=20, **kwargs):
        points = [
            x1 + radius, y1, x2 - radius, y1,
            x2, y1, x2, y1 + radius,
            x2, y2 - radius, x2, y2,
            x2 - radius, y2, x1 + radius, y2,
            x1, y2, x1, y2 - radius,
            x1, y1 + radius, x1, y1,
        ]
        return canvas.create_polygon(
            points, smooth=True, splinesteps=24, **kwargs
        )

    # 标题栏图标直接绘制，打包后无需附带额外 ico 文件。
    app_icon = tk.PhotoImage(width=32, height=32)
    app_icon.put('#5AA8FF', to=(5, 17, 10, 27))
    app_icon.put('#2F7BF3', to=(13, 10, 19, 27))
    app_icon.put('#1E5ED0', to=(22, 14, 27, 27))
    root.iconphoto(True, app_icon)

    canvas.create_text(
        48, 38, anchor='nw', text='营业额统计',
        font=('Segoe UI Variable Display', 26, 'bold'), fill=TEXT
    )
    canvas.create_text(
        50, 83, anchor='nw',
        text='合并钉钉记录与日流水，一键生成统计表',
        font=('Microsoft YaHei UI', 10), fill=MUTED
    )

    # 状态胶囊。
    rounded_rectangle(
        758, 44, 872, 84, radius=20,
        fill='#ECF7F0', outline='#FFFFFF', width=1
    )
    canvas.create_oval(
        774, 59, 784, 69, fill='#43B86A', outline='#2E9D53'
    )
    status_label = tk.Label(
        canvas, textvariable=status_var, bg='#ECF7F0', fg=SUCCESS,
        font=('Microsoft YaHei UI', 9, 'bold')
    )
    canvas.create_window(826, 64, window=status_label)

    # 高透明玻璃卡：减弱实色覆盖，让底部光晕更明显。
    rounded_rectangle(37, 132, 883, 489, radius=20,
                      fill='#C5D5E6', outline='')
    rounded_rectangle(34, 127, 880, 484, radius=20,
                      fill=CARD_BG, outline='#FFFFFF', width=1)
    rounded_rectangle(37, 508, 883, 687, radius=20,
                      fill='#C5D5E6', outline='')
    rounded_rectangle(34, 503, 880, 682, radius=20,
                      fill=CARD_BG, outline='#FFFFFF', width=1)

    form_card = tk.Frame(canvas, bg=CARD_BG, bd=0, highlightthickness=0)
    canvas.create_window(457, 306, window=form_card, width=800, height=320)
    log_card = tk.Frame(canvas, bg=CARD_BG, bd=0, highlightthickness=0)
    canvas.create_window(457, 592, window=log_card, width=800, height=140)

    style = ttk.Style(root)
    try:
        style.theme_use('clam')
    except tk.TclError:
        pass

    style.configure('Card.TFrame', background=CARD_BG)
    style.configure('Status.TLabel', background=CARD_BG, foreground=MUTED,
                    font=('Microsoft YaHei UI', 9))
    style.configure('Modern.TEntry', font=('Microsoft YaHei UI', 10),
                    padding=(8, 8), fieldbackground=FIELD_BG,
                    foreground=TEXT, borderwidth=0, relief='flat')
    style.map(
        'Modern.TEntry',
        fieldbackground=[('disabled', '#EDF1F6')],
        foreground=[('disabled', '#98A2B3')]
    )
    style.configure('Browse.TButton', font=('Microsoft YaHei UI', 9),
                    foreground=TEXT, background='#F8FAFD',
                    bordercolor=BORDER, borderwidth=1,
                    padding=(14, 8), relief='flat')
    style.map(
        'Browse.TButton',
        background=[('active', '#EAF2FF'), ('pressed', '#DCE9FF'),
                    ('disabled', '#EEF2F6')],
        foreground=[('disabled', '#98A2B3')]
    )
    style.configure('Accent.TButton', font=('Microsoft YaHei UI', 11, 'bold'),
                    foreground='#FFFFFF', background=PRIMARY,
                    bordercolor=PRIMARY, borderwidth=1,
                    padding=(42, 12), relief='flat')
    style.map(
        'Accent.TButton',
        background=[('active', PRIMARY_HOVER), ('pressed', '#1746B5'),
                    ('disabled', '#9DB7E8')],
        foreground=[('disabled', '#F6F9FF')]
    )
    style.configure('Subtle.TButton', font=('Microsoft YaHei UI', 9),
                    foreground=MUTED, background=CARD_BG,
                    borderwidth=0, padding=(10, 4), relief='flat')
    style.map(
        'Subtle.TButton',
        foreground=[('active', PRIMARY), ('pressed', '#1746B5')],
        background=[('active', '#E8F2FD')]
    )
    style.configure('Glass.Horizontal.TProgressbar', troughcolor='#DDE8F5',
                    background=PRIMARY, lightcolor=PRIMARY,
                    darkcolor=PRIMARY, borderwidth=0, thickness=4)

    def select_file(var, label, init_dir):
        path = filedialog.askopenfilename(
            title=f'选择{label}文件', initialdir=init_dir,
            filetypes=[('Excel 文件', '*.xlsx *.xls'), ('所有文件', '*.*')]
        )
        if path:
            var.set(path)

    def select_save_file(var, init_dir):
        path = filedialog.asksaveasfilename(
            title='选择输出文件位置', initialdir=init_dir,
            defaultextension='.xlsx', filetypes=[('Excel 文件', '*.xlsx')]
        )
        if path:
            var.set(path)

    form_card.columnconfigure(0, weight=1)
    form_card.columnconfigure(1, weight=1)
    field_entries = []
    browse_buttons = []

    def build_field(parent, row, column, label, variable, icon,
                    browse_command=None, helper=None):
        outer = tk.Frame(parent, bg=CARD_BG, bd=0)
        outer.grid(
            row=row, column=column, sticky='ew',
            padx=(12 if column == 0 else 18,
                  18 if column == 0 else 12),
            pady=(9, 7)
        )
        outer.columnconfigure(0, weight=1)

        tk.Label(
            outer, text=label, bg=CARD_BG, fg=TEXT,
            font=('Microsoft YaHei UI', 10, 'bold')
        ).grid(row=0, column=0, sticky='w', pady=(0, 7))

        field = tk.Frame(
            outer, bg=FIELD_BG, height=45, highlightthickness=1,
            highlightbackground=BORDER, highlightcolor=FOCUS, bd=0
        )
        field.grid(row=1, column=0, sticky='ew')
        field.grid_propagate(False)
        field.columnconfigure(1, weight=1)

        tk.Label(
            field, text=icon, bg=FIELD_BG, fg=PRIMARY,
            font=('Segoe Fluent Icons', 15)
        ).grid(row=0, column=0, padx=(12, 5), sticky='w')

        entry = ttk.Entry(field, textvariable=variable, style='Modern.TEntry')
        entry.grid(row=0, column=1, sticky='nsew')
        field_entries.append(entry)

        def focus_on(_event):
            field.configure(highlightbackground=FOCUS, highlightthickness=2)

        def focus_off(_event):
            field.configure(highlightbackground=BORDER, highlightthickness=1)

        entry.bind('<FocusIn>', focus_on)
        entry.bind('<FocusOut>', focus_off)

        if browse_command:
            button = ttk.Button(
                field, text='浏览', style='Browse.TButton',
                command=browse_command, takefocus=True
            )
            button.grid(row=0, column=2, sticky='ns', padx=4, pady=4)
            browse_buttons.append(button)

        if helper:
            tk.Label(
                outer, text=helper, bg=CARD_BG, fg=MUTED,
                font=('Microsoft YaHei UI', 8)
            ).grid(row=2, column=0, sticky='w', pady=(5, 0))
        return entry

    dd_entry = build_field(
        form_card, 0, 0, '钉钉记录', dd_var, '\ue8a5',
        lambda: select_file(dd_var, '钉钉记录', SCRIPT_DIR)
    )
    build_field(
        form_card, 0, 1, '日流水表', rl_var, '\ue8a5',
        lambda: select_file(rl_var, '日流水表', SCRIPT_DIR)
    )
    build_field(
        form_card, 1, 0, '统计日期', date_var, '\ue787',
        helper='留空将自动从钉钉记录提取'
    )
    build_field(
        form_card, 1, 1, '输出位置', output_var, '\ue8b7',
        lambda: select_save_file(output_var, SCRIPT_DIR)
    )

    action_inner = ttk.Frame(form_card, style='Card.TFrame')
    action_inner.grid(row=2, column=0, columnspan=2, pady=(12, 0))
    gen_btn = ttk.Button(
        action_inner, text='生成统计表', style='Accent.TButton',
        command=lambda: on_generate()
    )
    gen_btn.grid(row=0, column=0, rowspan=2)
    ttk.Label(
        action_inner, textvariable=action_status_var, style='Status.TLabel'
    ).grid(row=0, column=1, sticky='sw', padx=(16, 0))
    progress = ttk.Progressbar(
        action_inner, mode='indeterminate', length=136,
        style='Glass.Horizontal.TProgressbar'
    )
    progress.grid(row=1, column=1, sticky='nw', padx=(16, 0), pady=(4, 0))
    progress.grid_remove()

    log_header = tk.Frame(log_card, bg=CARD_BG)
    log_header.pack(fill='x', padx=4, pady=(0, 7))
    tk.Label(
        log_header, text='运行日志', bg=CARD_BG, fg=TEXT,
        font=('Microsoft YaHei UI', 10, 'bold')
    ).pack(side='left')

    log_body = tk.Frame(
        log_card, bg=LOG_BG, highlightthickness=1,
        highlightbackground='#D5DFEB', bd=0
    )
    log_body.pack(fill='both', expand=True, padx=2, pady=(0, 2))
    log_text = tk.Text(
        log_body, font=('Cascadia Mono', 9), height=5, wrap='word',
        bg=LOG_BG, fg='#455269', insertbackground=TEXT,
        selectbackground='#CFE3FF', relief='flat', bd=0,
        padx=12, pady=9, highlightthickness=0
    )
    log_scroll = ttk.Scrollbar(log_body, command=log_text.yview)
    log_text.configure(yscrollcommand=log_scroll.set)
    log_scroll.pack(side='right', fill='y')
    log_text.pack(side='left', fill='both', expand=True)
    log_text.tag_configure('error', foreground='#C93645')
    log_text.tag_configure(
        'alert', foreground='#C62828',
        font=('Cascadia Mono', 9, 'bold')
    )
    log_text.tag_configure('success', foreground='#248A4B')
    log_text.insert(
        'end', '系统已就绪\n请选择数据文件\n生成结果将在此处显示\n'
    )
    log_text.configure(state='disabled')

    def clear_log():
        log_text.configure(state='normal')
        log_text.delete('1.0', 'end')
        log_text.configure(state='disabled')

    clear_btn = ttk.Button(
        log_header, text='清空', style='Subtle.TButton', command=clear_log
    )
    clear_btn.pack(side='right')
    status_dot = tk.Canvas(
        log_header, width=14, height=14, bg=CARD_BG, highlightthickness=0
    )
    status_dot.create_oval(3, 3, 11, 11, fill='#43B86A', outline='#2E9D53')
    status_dot.pack(side='right', padx=(0, 8))

    def append_log(msg):
        msg = str(msg)
        tag = ''
        if msg.startswith('[异常提醒]'):
            tag = 'alert'
        elif '错误' in msg or '失败' in msg:
            tag = 'error'
        elif '完成' in msg:
            tag = 'success'
        log_text.configure(state='normal')
        log_text.insert('end', msg + '\n', tag)
        log_text.see('end')
        log_text.configure(state='disabled')

    def log(msg):
        log_queue.put(msg)

    def poll_log_queue():
        try:
            while True:
                append_log(log_queue.get_nowait())
        except queue.Empty:
            pass
        root.after(80, poll_log_queue)

    def show_success(out, n, regions):
        output_var.set(out)
        status_var.set('生成完成')
        action_status_var.set(f'完成 · {n} 家门店 / {regions} 个区域')
        messagebox.showinfo(
            '生成完成',
            f'统计表已生成：\n{out}\n\n共 {n} 家门店，{regions} 个区域'
        )

    def show_failure(message):
        status_var.set('生成失败')
        action_status_var.set('请查看运行日志')
        messagebox.showerror('生成失败', message)

    def finish_generation():
        gen_btn.configure(state='normal', text='生成统计表')
        for widget in field_entries + browse_buttons:
            widget.configure(state='normal')
        clear_btn.configure(state='normal')
        progress.stop()
        progress.grid_remove()
        generating['state'] = False

    def on_generate():
        if generating['state']:
            return
        dd_path = dd_var.get().strip()
        rl_path = rl_var.get().strip()
        out_path = output_var.get().strip() or None
        d_str = date_var.get().strip() or None

        if not dd_path:
            messagebox.showwarning('提示', '请先选择钉钉记录文件')
            return
        if not os.path.exists(dd_path):
            messagebox.showerror('错误', f'钉钉记录文件不存在：\n{dd_path}')
            return
        if not rl_path:
            messagebox.showwarning('提示', '请先选择日流水文件')
            return
        if not os.path.exists(rl_path):
            messagebox.showerror('错误', f'日流水文件不存在：\n{rl_path}')
            return

        clear_log()
        gen_btn.configure(state='disabled', text='正在生成…')
        for widget in field_entries + browse_buttons:
            widget.configure(state='disabled')
        clear_btn.configure(state='disabled')
        status_var.set('正在处理')
        action_status_var.set('正在读取并合并数据…')
        progress.grid()
        progress.start(12)
        generating['state'] = True

        def task():
            try:
                result = run_pipeline(
                    dd_path, rl_path, out_path, d_str, log=log
                )
                out, n, regions = result
                if out:
                    root.after(
                        0, lambda out=out, n=n, regions=regions:
                        show_success(out, n, regions)
                    )
                else:
                    root.after(
                        0, lambda: show_failure('生成失败，请查看运行日志')
                    )
            except Exception as error:
                log(f'\n错误: {error}')
                root.after(0, lambda err=error: show_failure(str(err)))
            finally:
                root.after(0, finish_generation)

        threading.Thread(target=task, daemon=True).start()

    dd_entry.focus_set()
    root.bind('<Control-Return>', lambda _event: on_generate())
    root.bind(
        '<Escape>',
        lambda _event: root.destroy() if not generating['state'] else None
    )
    root.after(80, poll_log_queue)
    root.mainloop()


def gui_main_clean():
    """更克制的 Win11 Fluent 界面：背景有磨砂感，内容层保持清晰。"""
    import tkinter as tk
    from tkinter import ttk, filedialog, messagebox
    import threading
    import queue
    import ctypes

    if sys.platform == 'win32':
        try:
            ctypes.windll.user32.SetProcessDpiAwarenessContext(ctypes.c_void_p(-4))
        except (AttributeError, OSError):
            try:
                ctypes.windll.shcore.SetProcessDpiAwareness(2)
            except (AttributeError, OSError):
                pass

    root = tk.Tk()
    root.title('营业额统计')
    root.resizable(False, False)

    window_width = 840
    window_height = 680
    x = (root.winfo_screenwidth() - window_width) // 2
    y = (root.winfo_screenheight() - window_height) // 2
    root.geometry(f'{window_width}x{window_height}+{x}+{y}')

    # 不再对整个窗口应用 alpha：避免桌面图标和文字穿透控件。
    try:
        root.attributes('-alpha', 1.0)
    except tk.TclError:
        pass

    def enable_windows_11_effects():
        if sys.platform != 'win32':
            return
        try:
            root.update_idletasks()
            hwnd = ctypes.windll.user32.GetParent(root.winfo_id())
            corner_preference = ctypes.c_int(2)
            backdrop_type = ctypes.c_int(2)
            ctypes.windll.dwmapi.DwmSetWindowAttribute(
                hwnd, 33, ctypes.byref(corner_preference),
                ctypes.sizeof(corner_preference)
            )
            ctypes.windll.dwmapi.DwmSetWindowAttribute(
                hwnd, 38, ctypes.byref(backdrop_type),
                ctypes.sizeof(backdrop_type)
            )
        except (AttributeError, OSError):
            pass

    root.after(0, enable_windows_11_effects)

    dd_var = tk.StringVar(value=os.path.join(SCRIPT_DIR, '钉钉记录.xlsx'))
    rl_var = tk.StringVar(value=os.path.join(SCRIPT_DIR, '日流水.xls'))
    date_var = tk.StringVar()
    output_var = tk.StringVar()
    status_var = tk.StringVar(value='准备就绪')
    action_status_var = tk.StringVar(value='等待开始')
    generating = {'state': False}
    log_queue = queue.Queue()

    BG_TOP = '#EAF4FF'
    BG_BOTTOM = '#F2EEFF'
    SURFACE = '#F8FBFF'
    SURFACE_ALT = '#F3F7FC'
    FIELD = '#FFFFFF'
    TEXT = '#172033'
    MUTED = '#667085'
    BORDER = '#D5DFEB'
    PRIMARY = '#2563EB'
    PRIMARY_HOVER = '#1D4ED8'
    FOCUS = '#3B82F6'
    SUCCESS = '#21884A'

    canvas = tk.Canvas(
        root, width=window_width, height=window_height,
        highlightthickness=0, bd=0, bg=BG_TOP
    )
    canvas.pack(fill='both', expand=True)

    def hex_to_rgb(color):
        color = color.lstrip('#')
        return tuple(int(color[i:i + 2], 16) for i in (0, 2, 4))

    def mix_color(start, end, ratio):
        a = hex_to_rgb(start)
        b = hex_to_rgb(end)
        values = [round(a[i] + (b[i] - a[i]) * ratio) for i in range(3)]
        return '#{:02x}{:02x}{:02x}'.format(*values)

    for line_y in range(0, window_height, 2):
        ratio = line_y / max(window_height - 1, 1)
        canvas.create_rectangle(
            0, line_y, window_width, line_y + 2,
            fill=mix_color(BG_TOP, BG_BOTTOM, ratio), outline=''
        )

    # 仅在背景中保留两块柔和色彩，模拟 Mica 氛围，不影响内容可读性。
    canvas.create_oval(-170, 410, 280, 860, fill='#E5DEFF', outline='')
    canvas.create_oval(650, 410, 1080, 840, fill='#DDF6FF', outline='')

    def rounded_rectangle(x1, y1, x2, y2, radius=18, **kwargs):
        points = [
            x1 + radius, y1, x2 - radius, y1,
            x2, y1, x2, y1 + radius,
            x2, y2 - radius, x2, y2,
            x2 - radius, y2, x1 + radius, y2,
            x1, y2, x1, y2 - radius,
            x1, y1 + radius, x1, y1,
        ]
        return canvas.create_polygon(
            points, smooth=True, splinesteps=24, **kwargs
        )

    app_icon = tk.PhotoImage(width=32, height=32)
    app_icon.put('#5AA8FF', to=(5, 17, 10, 27))
    app_icon.put('#2F7BF3', to=(13, 10, 19, 27))
    app_icon.put('#1E5ED0', to=(22, 14, 27, 27))
    root.iconphoto(True, app_icon)

    canvas.create_text(
        38, 28, anchor='nw', text='营业额统计',
        font=('Segoe UI Variable Display', 23, 'bold'), fill=TEXT
    )
    canvas.create_text(
        40, 68, anchor='nw', text='合并钉钉记录与日流水，生成统一统计表',
        font=('Microsoft YaHei UI', 9), fill=MUTED
    )

    rounded_rectangle(
        690, 33, 802, 73, radius=20,
        fill='#EFF8F2', outline='#FFFFFF', width=1
    )
    canvas.create_oval(706, 48, 716, 58, fill='#43B86A', outline='#2E9D53')
    status_label = tk.Label(
        canvas, textvariable=status_var, bg='#EFF8F2', fg=SUCCESS,
        font=('Microsoft YaHei UI', 9, 'bold')
    )
    canvas.create_window(757, 53, window=status_label)

    # 阴影只偏移 3px，卡片高度随内容收紧。
    rounded_rectangle(29, 112, 813, 432, radius=18,
                      fill='#CBD8E6', outline='')
    rounded_rectangle(26, 108, 810, 428, radius=18,
                      fill=SURFACE, outline='#FFFFFF', width=1)
    rounded_rectangle(29, 457, 813, 652, radius=18,
                      fill='#CBD8E6', outline='')
    rounded_rectangle(26, 453, 810, 648, radius=18,
                      fill=SURFACE, outline='#FFFFFF', width=1)

    form_card = tk.Frame(canvas, bg=SURFACE, bd=0, highlightthickness=0)
    canvas.create_window(418, 268, window=form_card, width=720, height=270)
    log_card = tk.Frame(canvas, bg=SURFACE, bd=0, highlightthickness=0)
    canvas.create_window(418, 550, window=log_card, width=720, height=150)

    style = ttk.Style(root)
    try:
        style.theme_use('clam')
    except tk.TclError:
        pass

    style.configure(
        'Clean.TEntry', font=('Microsoft YaHei UI', 9),
        padding=(10, 8), fieldbackground=FIELD, foreground=TEXT,
        bordercolor=BORDER, lightcolor=BORDER, darkcolor=BORDER,
        borderwidth=1, relief='flat'
    )
    style.map(
        'Clean.TEntry',
        bordercolor=[('focus', FOCUS)],
        lightcolor=[('focus', FOCUS)],
        darkcolor=[('focus', FOCUS)],
        fieldbackground=[('disabled', '#EEF2F6')],
        foreground=[('disabled', '#98A2B3')]
    )
    style.configure(
        'CleanBrowse.TButton', font=('Microsoft YaHei UI', 9),
        foreground=TEXT, background='#F7F9FC', bordercolor=BORDER,
        borderwidth=1, padding=(14, 8), relief='flat'
    )
    style.map(
        'CleanBrowse.TButton',
        background=[('active', '#EAF2FF'), ('pressed', '#DCE9FF'),
                    ('disabled', '#EEF2F6')],
        foreground=[('disabled', '#98A2B3')]
    )
    style.configure(
        'CleanAccent.TButton', font=('Microsoft YaHei UI', 10, 'bold'),
        foreground='#FFFFFF', background=PRIMARY, bordercolor=PRIMARY,
        borderwidth=1, padding=(30, 10), relief='flat'
    )
    style.map(
        'CleanAccent.TButton',
        background=[('active', PRIMARY_HOVER), ('pressed', '#1746B5'),
                    ('disabled', '#9DB7E8')],
        foreground=[('disabled', '#F6F9FF')]
    )
    style.configure(
        'CleanSubtle.TButton', font=('Microsoft YaHei UI', 9),
        foreground=MUTED, background=SURFACE, borderwidth=0,
        padding=(9, 4), relief='flat'
    )
    style.map(
        'CleanSubtle.TButton',
        foreground=[('active', PRIMARY)],
        background=[('active', '#EAF2FC')]
    )
    style.configure(
        'Clean.Horizontal.TProgressbar', troughcolor='#DDE8F5',
        background=PRIMARY, lightcolor=PRIMARY, darkcolor=PRIMARY,
        borderwidth=0, thickness=4
    )

    def select_file(var, label):
        path = filedialog.askopenfilename(
            title=f'选择{label}文件', initialdir=SCRIPT_DIR,
            filetypes=[('Excel 文件', '*.xlsx *.xls'), ('所有文件', '*.*')]
        )
        if path:
            var.set(path)

    def select_save_file():
        path = filedialog.asksaveasfilename(
            title='选择输出文件位置', initialdir=SCRIPT_DIR,
            defaultextension='.xlsx', filetypes=[('Excel 文件', '*.xlsx')]
        )
        if path:
            output_var.set(path)

    tk.Label(
        form_card, text='数据配置', bg=SURFACE, fg=TEXT,
        font=('Microsoft YaHei UI', 11, 'bold')
    ).grid(row=0, column=0, columnspan=3, sticky='w', pady=(0, 12))

    form_card.columnconfigure(0, minsize=92)
    form_card.columnconfigure(1, weight=1)
    form_card.columnconfigure(2, minsize=80)
    field_entries = []
    browse_buttons = []

    def add_row(row, label, variable, browse_command=None, trailing_text=''):
        tk.Label(
            form_card, text=label, bg=SURFACE, fg=TEXT,
            font=('Microsoft YaHei UI', 9, 'bold')
        ).grid(row=row, column=0, sticky='w', pady=5)
        entry = ttk.Entry(
            form_card, textvariable=variable, style='Clean.TEntry'
        )
        entry.grid(row=row, column=1, sticky='ew', pady=5, padx=(0, 10))
        field_entries.append(entry)
        if browse_command:
            button = ttk.Button(
                form_card, text='浏览', style='CleanBrowse.TButton',
                command=browse_command, takefocus=True
            )
            button.grid(row=row, column=2, sticky='ew', pady=5)
            browse_buttons.append(button)
        else:
            tk.Label(
                form_card, text=trailing_text, bg=SURFACE, fg=MUTED,
                font=('Microsoft YaHei UI', 8)
            ).grid(row=row, column=2, sticky='w', pady=5)
        return entry

    dd_entry = add_row(
        1, '钉钉记录', dd_var,
        lambda: select_file(dd_var, '钉钉记录')
    )
    add_row(
        2, '日流水表', rl_var,
        lambda: select_file(rl_var, '日流水表')
    )
    add_row(3, '统计日期', date_var, trailing_text='自动提取')
    add_row(4, '输出位置', output_var, select_save_file)

    action_row = tk.Frame(form_card, bg=SURFACE)
    action_row.grid(row=5, column=0, columnspan=3, sticky='ew', pady=(13, 0))
    action_row.columnconfigure(0, weight=1)
    action_status_label = tk.Label(
        action_row, textvariable=action_status_var, bg=SURFACE, fg=MUTED,
        font=('Microsoft YaHei UI', 9)
    )
    action_status_label.grid(row=0, column=0, sticky='w')
    progress = ttk.Progressbar(
        action_row, mode='indeterminate', length=150,
        style='Clean.Horizontal.TProgressbar'
    )
    progress.grid(row=1, column=0, sticky='w', pady=(5, 0))
    progress.grid_remove()
    gen_btn = ttk.Button(
        action_row, text='生成统计表', style='CleanAccent.TButton',
        command=lambda: on_generate()
    )
    gen_btn.grid(row=0, column=1, rowspan=2, sticky='e')

    log_header = tk.Frame(log_card, bg=SURFACE)
    log_header.pack(fill='x', pady=(0, 7))
    tk.Label(
        log_header, text='运行日志', bg=SURFACE, fg=TEXT,
        font=('Microsoft YaHei UI', 10, 'bold')
    ).pack(side='left')

    log_body = tk.Frame(
        log_card, bg=SURFACE_ALT, highlightthickness=1,
        highlightbackground=BORDER, bd=0
    )
    log_body.pack(fill='both', expand=True)
    log_text = tk.Text(
        log_body, font=('Cascadia Mono', 9), height=5, wrap='word',
        bg=SURFACE_ALT, fg='#455269', insertbackground=TEXT,
        selectbackground='#CFE3FF', relief='flat', bd=0,
        padx=11, pady=8, highlightthickness=0
    )
    log_scroll = ttk.Scrollbar(log_body, command=log_text.yview)
    log_text.configure(yscrollcommand=log_scroll.set)
    log_scroll.pack(side='right', fill='y')
    log_text.pack(side='left', fill='both', expand=True)
    log_text.tag_configure('error', foreground='#C93645')
    log_text.tag_configure(
        'alert', foreground='#C62828', font=('Cascadia Mono', 9, 'bold')
    )
    log_text.tag_configure('success', foreground=SUCCESS)
    log_text.insert('end', '系统已就绪\n请选择数据文件\n')
    log_text.configure(state='disabled')

    def clear_log():
        log_text.configure(state='normal')
        log_text.delete('1.0', 'end')
        log_text.configure(state='disabled')

    clear_btn = ttk.Button(
        log_header, text='清空', style='CleanSubtle.TButton', command=clear_log
    )
    clear_btn.pack(side='right')
    dot = tk.Canvas(
        log_header, width=14, height=14, bg=SURFACE, highlightthickness=0
    )
    dot.create_oval(3, 3, 11, 11, fill='#43B86A', outline='#2E9D53')
    dot.pack(side='right', padx=(0, 8))

    def append_log(msg):
        msg = str(msg)
        tag = ''
        if msg.startswith('[异常提醒]'):
            tag = 'alert'
        elif '错误' in msg or '失败' in msg:
            tag = 'error'
        elif '完成' in msg:
            tag = 'success'
        log_text.configure(state='normal')
        log_text.insert('end', msg + '\n', tag)
        log_text.see('end')
        log_text.configure(state='disabled')

    def log(msg):
        log_queue.put(msg)

    def poll_log_queue():
        try:
            while True:
                append_log(log_queue.get_nowait())
        except queue.Empty:
            pass
        root.after(80, poll_log_queue)

    def show_success(out, n, regions):
        output_var.set(out)
        status_var.set('生成完成')
        action_status_var.set(f'完成 · {n} 家门店 / {regions} 个区域')
        messagebox.showinfo(
            '生成完成',
            f'统计表已生成：\n{out}\n\n共 {n} 家门店，{regions} 个区域'
        )

    def show_failure(message):
        status_var.set('生成失败')
        action_status_var.set('请查看运行日志')
        messagebox.showerror('生成失败', message)

    def finish_generation():
        gen_btn.configure(state='normal', text='生成统计表')
        for widget in field_entries + browse_buttons:
            widget.configure(state='normal')
        clear_btn.configure(state='normal')
        progress.stop()
        progress.grid_remove()
        generating['state'] = False

    def on_generate():
        if generating['state']:
            return
        dd_path = dd_var.get().strip()
        rl_path = rl_var.get().strip()
        out_path = output_var.get().strip() or None
        d_str = date_var.get().strip() or None

        if not dd_path:
            messagebox.showwarning('提示', '请先选择钉钉记录文件')
            return
        if not os.path.exists(dd_path):
            messagebox.showerror('错误', f'钉钉记录文件不存在：\n{dd_path}')
            return
        if not rl_path:
            messagebox.showwarning('提示', '请先选择日流水文件')
            return
        if not os.path.exists(rl_path):
            messagebox.showerror('错误', f'日流水文件不存在：\n{rl_path}')
            return

        clear_log()
        gen_btn.configure(state='disabled', text='正在生成…')
        for widget in field_entries + browse_buttons:
            widget.configure(state='disabled')
        clear_btn.configure(state='disabled')
        status_var.set('正在处理')
        action_status_var.set('正在读取并合并数据…')
        progress.grid()
        progress.start(12)
        generating['state'] = True

        def task():
            try:
                out, n, regions = run_pipeline(
                    dd_path, rl_path, out_path, d_str, log=log
                )
                if out:
                    root.after(
                        0, lambda out=out, n=n, regions=regions:
                        show_success(out, n, regions)
                    )
                else:
                    root.after(
                        0, lambda: show_failure('生成失败，请查看运行日志')
                    )
            except Exception as error:
                log(f'\n错误: {error}')
                root.after(0, lambda err=error: show_failure(str(err)))
            finally:
                root.after(0, finish_generation)

        threading.Thread(target=task, daemon=True).start()

    dd_entry.focus_set()
    root.bind('<Control-Return>', lambda _event: on_generate())
    root.bind(
        '<Escape>',
        lambda _event: root.destroy() if not generating['state'] else None
    )
    root.after(80, poll_log_queue)
    root.mainloop()


def gui_main_reference():
    """参考居中工作台方案的 Win11 界面。"""
    import tkinter as tk
    from tkinter import ttk, filedialog, messagebox
    import threading
    import queue
    import ctypes

    if sys.platform == 'win32':
        try:
            ctypes.windll.user32.SetProcessDpiAwarenessContext(ctypes.c_void_p(-4))
        except (AttributeError, OSError):
            try:
                ctypes.windll.shcore.SetProcessDpiAwareness(2)
            except (AttributeError, OSError):
                pass

    root = tk.Tk()
    root.title('营业额统计')
    root.resizable(True, False)
    window_width, window_height = 1120, 720
    root.minsize(980, window_height)
    x = (root.winfo_screenwidth() - window_width) // 2
    y = (root.winfo_screenheight() - window_height) // 2
    root.geometry(f'{window_width}x{window_height}+{x}+{y}')

    def enable_windows_11_effects():
        if sys.platform != 'win32':
            return
        try:
            root.update_idletasks()
            hwnd = ctypes.windll.user32.GetParent(root.winfo_id())
            corner = ctypes.c_int(2)
            mica = ctypes.c_int(2)
            ctypes.windll.dwmapi.DwmSetWindowAttribute(
                hwnd, 33, ctypes.byref(corner), ctypes.sizeof(corner)
            )
            ctypes.windll.dwmapi.DwmSetWindowAttribute(
                hwnd, 38, ctypes.byref(mica), ctypes.sizeof(mica)
            )
        except (AttributeError, OSError):
            pass

    root.after(0, enable_windows_11_effects)

    dd_var = tk.StringVar(value=os.path.join(SCRIPT_DIR, '钉钉记录.xlsx'))
    rl_var = tk.StringVar(value=os.path.join(SCRIPT_DIR, '日流水.xls'))
    date_var = tk.StringVar()
    output_var = tk.StringVar()
    status_var = tk.StringVar(value='等待文件')
    action_status_var = tk.StringVar(value='请选择两个 Excel 文件开始处理')
    dd_display = tk.StringVar()
    rl_display = tk.StringVar()
    generating = {'state': False}
    log_queue = queue.Queue()

    BG = '#EDF9F8'
    GRID = '#DFEFEE'
    SURFACE = '#F9FCFC'
    SURFACE_ALT = '#F3F9F9'
    WHITE = '#FFFFFF'
    TEXT = '#123F3D'
    MUTED = '#688684'
    TEAL = '#119D96'
    TEAL_DARK = '#087D78'
    BLUE = '#2D7DD2'
    BORDER = '#CFE4E2'
    SUCCESS = '#218B65'

    canvas = tk.Canvas(
        root, width=window_width, height=window_height,
        bg=BG, highlightthickness=0, bd=0
    )
    canvas.pack(fill='both', expand=True)

    # 极浅网格和边缘光斑，只服务于氛围，不穿透内容层。
    for grid_x in range(0, root.winfo_screenwidth() + 28, 28):
        canvas.create_line(grid_x, 0, grid_x, window_height, fill=GRID)
    for grid_y in range(0, window_height, 28):
        canvas.create_line(0, grid_y, window_width, grid_y, fill=GRID)
    canvas.create_oval(-170, -150, 170, 190, fill='#C8F5EF', outline='')
    right_glow = canvas.create_oval(
        window_width - 175, 185, window_width + 150, 510,
        fill='#D7E8FF', outline=''
    )

    def rounded_points(x1, y1, x2, y2, radius=18):
        return [
            x1 + radius, y1, x2 - radius, y1,
            x2, y1, x2, y1 + radius,
            x2, y2 - radius, x2, y2,
            x2 - radius, y2, x1 + radius, y2,
            x1, y2, x1, y2 - radius,
            x1, y1 + radius, x1, y1,
        ]

    def rounded_rectangle(x1, y1, x2, y2, radius=18, **kwargs):
        return canvas.create_polygon(
            rounded_points(x1, y1, x2, y2, radius),
            smooth=True, splinesteps=24, **kwargs
        )

    try:
        root.iconbitmap(APP_ICON_PATH)
    except tk.TclError:
        pass

    # 顶部品牌栏。
    top_bar = rounded_rectangle(
        120, 16, window_width - 120, 66, radius=18,
        fill='#F8FEFD', outline='#FFFFFF', width=1
    )
    canvas.create_oval(134, 26, 168, 60, fill=TEAL, outline='')
    canvas.create_text(
        151, 43, text='营', font=('Microsoft YaHei UI', 11, 'bold'),
        fill=WHITE
    )
    canvas.create_text(
        180, 29, anchor='nw', text='营业额统计',
        font=('Microsoft YaHei UI', 10, 'bold'), fill=TEXT
    )
    canvas.create_text(
        180, 47, anchor='nw', text='移动数据工作台',
        font=('Microsoft YaHei UI', 8, 'bold'), fill=MUTED
    )
    center_brand = canvas.create_text(
        window_width / 2, 41, text='X-ME  |  晴',
        font=('Segoe UI Variable', 10, 'bold'), fill=TEAL_DARK
    )
    status_surface = rounded_rectangle(
        window_width - 235, 26, window_width - 136, 57, radius=15,
        fill='#E5F7F5', outline='#C8EAE7', width=1
    )
    status_dot = canvas.create_oval(
        window_width - 223, 38, window_width - 215, 46,
        fill=TEAL, outline=''
    )
    top_status = tk.Label(
        canvas, textvariable=status_var, bg='#E5F7F5', fg=TEAL_DARK,
        font=('Microsoft YaHei UI', 8, 'bold')
    )
    top_status_window = canvas.create_window(
        window_width - 178, 42, window=top_status
    )

    # 居中标题与功能标签。
    hero_title = canvas.create_text(
        window_width / 2, 92, text='两份流水  一键合并',
        font=('Microsoft YaHei UI', 25, 'bold'), fill=TEXT
    )
    tag_specs = [('自动匹配', -68), ('异常标红', 0), ('高清导出', 68)]
    tag_items = []
    for tag_text, tag_offset in tag_specs:
        tag_x = window_width / 2 + tag_offset
        tag_surface = rounded_rectangle(
            tag_x - 29, 119, tag_x + 29, 143, radius=12,
            fill='#F8FEFD', outline='#FFFFFF', width=1
        )
        tag_label = canvas.create_text(
            tag_x, 131, text=tag_text,
            font=('Microsoft YaHei UI', 8, 'bold'), fill=MUTED
        )
        tag_items.append((tag_surface, tag_label, tag_offset))

    # 中央工作台。
    main_shadow = rounded_rectangle(
        105, 165, window_width - 95, 687, radius=23,
        fill='#C9DCDA', outline=''
    )
    main_surface = rounded_rectangle(
        100, 160, window_width - 100, 682, radius=23,
        fill=SURFACE, outline='#FFFFFF', width=1
    )
    workbench = tk.Frame(canvas, bg=SURFACE, bd=0, highlightthickness=0)
    workbench_window = canvas.create_window(
        window_width / 2, 421, window=workbench,
        width=window_width - 260, height=480
    )

    def resize_layout(event):
        """保持内容居中，并让表单随窗口宽度同步伸缩。"""
        width = max(event.width, 980)
        center = width / 2
        canvas.coords(right_glow, width - 175, 185, width + 150, 510)
        canvas.coords(
            top_bar, *rounded_points(120, 16, width - 120, 66, 18)
        )
        canvas.coords(center_brand, center, 41)
        canvas.coords(
            status_surface,
            *rounded_points(width - 235, 26, width - 136, 57, 15)
        )
        canvas.coords(status_dot, width - 223, 38, width - 215, 46)
        canvas.coords(top_status_window, width - 178, 42)
        canvas.coords(hero_title, center, 92)
        for tag_surface, tag_label, tag_offset in tag_items:
            tag_x = center + tag_offset
            canvas.coords(
                tag_surface,
                *rounded_points(tag_x - 29, 119, tag_x + 29, 143, 12)
            )
            canvas.coords(tag_label, tag_x, 131)
        canvas.coords(
            main_shadow, *rounded_points(105, 165, width - 95, 687, 23)
        )
        canvas.coords(
            main_surface, *rounded_points(100, 160, width - 100, 682, 23)
        )
        canvas.coords(workbench_window, center, 421)
        canvas.itemconfigure(workbench_window, width=width - 260)

    canvas.bind('<Configure>', resize_layout)

    style = ttk.Style(root)
    try:
        style.theme_use('clam')
    except tk.TclError:
        pass
    style.configure(
        'Reference.TEntry', font=('Microsoft YaHei UI', 10, 'bold'),
        padding=(11, 9), fieldbackground=WHITE, foreground=TEXT,
        bordercolor=BORDER, lightcolor=BORDER, darkcolor=BORDER,
        borderwidth=1, relief='flat'
    )
    style.map(
        'Reference.TEntry',
        bordercolor=[('focus', TEAL)],
        lightcolor=[('focus', TEAL)],
        darkcolor=[('focus', TEAL)],
        fieldbackground=[('disabled', '#EDF4F3')],
        foreground=[('disabled', '#91A6A4')]
    )
    style.configure(
        'Reference.Horizontal.TProgressbar', troughcolor='#DDECEA',
        background=TEAL, lightcolor=TEAL, darkcolor=TEAL,
        borderwidth=0, thickness=4
    )

    def display_path(path):
        if not path:
            return '尚未选择文件'
        return os.path.basename(path)

    def refresh_display(*_args):
        dd_display.set(display_path(dd_var.get()))
        rl_display.set(display_path(rl_var.get()))

    dd_var.trace_add('write', refresh_display)
    rl_var.trace_add('write', refresh_display)
    refresh_display()

    def select_file(var, label):
        path = filedialog.askopenfilename(
            title=f'选择{label}文件', initialdir=SCRIPT_DIR,
            filetypes=[('Excel 文件', '*.xlsx *.xls'), ('所有文件', '*.*')]
        )
        if path:
            var.set(path)

    def select_save_file():
        path = filedialog.asksaveasfilename(
            title='选择输出文件位置', initialdir=SCRIPT_DIR,
            defaultextension='.xlsx', filetypes=[('Excel 文件', '*.xlsx')]
        )
        if path:
            output_var.set(path)

    # 工作台标题。
    heading = tk.Frame(workbench, bg=SURFACE, height=38)
    heading.pack(fill='x')
    heading.pack_propagate(False)
    step_badge = tk.Label(
        heading, text='01', bg=TEAL, fg=WHITE,
        font=('Segoe UI Variable', 9, 'bold'), padx=10, pady=6
    )
    step_badge.pack(side='left')
    tk.Label(
        heading, text='选择数据文件', bg=SURFACE, fg=TEXT,
        font=('Microsoft YaHei UI', 11, 'bold')
    ).pack(side='left', padx=(10, 0))
    tk.Label(
        heading, textvariable=status_var, bg=SURFACE_ALT, fg=MUTED,
        font=('Microsoft YaHei UI', 8, 'bold'), padx=9, pady=4
    ).pack(side='right')

    tiles = tk.Frame(workbench, bg=SURFACE)
    tiles.pack(fill='x', pady=(9, 9))
    tiles.columnconfigure(0, weight=1)
    tiles.columnconfigure(1, weight=1)
    tile_buttons = []

    def file_tile(parent, column, icon_text, title, display_var, command):
        tile = tk.Frame(
            parent, bg=WHITE, height=66, highlightthickness=1,
            highlightbackground=BORDER, bd=0
        )
        tile.grid(
            row=0, column=column, sticky='ew',
            padx=(0, 6) if column == 0 else (6, 0)
        )
        tile.grid_propagate(False)
        icon = tk.Label(
            tile, text=icon_text, bg='#E2F5F3', fg=TEAL_DARK,
            font=('Microsoft YaHei UI', 11, 'bold'), width=3, height=2
        )
        icon.pack(side='left', padx=(10, 9))
        copy = tk.Frame(tile, bg=WHITE)
        copy.pack(side='left', fill='both', expand=True, pady=10)
        tk.Label(
            copy, text=title, bg=WHITE, fg=TEXT,
            font=('Microsoft YaHei UI', 9, 'bold')
        ).pack(anchor='w')
        tk.Label(
            copy, textvariable=display_var, bg=WHITE, fg=MUTED,
            font=('Microsoft YaHei UI', 8, 'bold')
        ).pack(anchor='w', pady=(3, 0))
        button = tk.Button(
            tile, text='选择', command=command, bg=WHITE, fg=TEAL_DARK,
            activebackground='#E7F7F5', activeforeground=TEAL_DARK,
            font=('Microsoft YaHei UI', 8, 'bold'), relief='flat',
            bd=0, cursor='hand2', padx=10, pady=16, takefocus=True
        )
        button.pack(side='right', fill='y', padx=(2, 4), pady=4)
        tile_buttons.append(button)

    file_tile(
        tiles, 0, '钉', '钉钉记录', dd_display,
        lambda: select_file(dd_var, '钉钉记录')
    )
    file_tile(
        tiles, 1, '流', '日流水表', rl_display,
        lambda: select_file(rl_var, '日流水表')
    )

    fields = tk.Frame(workbench, bg=SURFACE)
    fields.pack(fill='x')
    fields.columnconfigure(0, weight=1)

    tk.Label(
        fields, text='统计日期', bg=SURFACE, fg=TEXT,
        font=('Microsoft YaHei UI', 8, 'bold')
    ).grid(row=0, column=0, sticky='w')
    tk.Label(
        fields, text='选填 · 留空自动提取', bg=SURFACE, fg=MUTED,
        font=('Microsoft YaHei UI', 8, 'bold')
    ).grid(row=0, column=0, sticky='e')
    date_entry = ttk.Entry(
        fields, textvariable=date_var, style='Reference.TEntry'
    )
    date_entry.grid(row=1, column=0, sticky='ew', pady=(5, 9))

    tk.Label(
        fields, text='输出位置', bg=SURFACE, fg=TEXT,
        font=('Microsoft YaHei UI', 8, 'bold')
    ).grid(row=2, column=0, sticky='w')
    output_row = tk.Frame(fields, bg=SURFACE)
    output_row.grid(row=3, column=0, sticky='ew', pady=(5, 10))
    output_row.columnconfigure(0, weight=1)
    output_entry = ttk.Entry(
        output_row, textvariable=output_var, style='Reference.TEntry'
    )
    output_entry.grid(row=0, column=0, sticky='ew', padx=(0, 8))
    output_button = tk.Button(
        output_row, text='选择位置', command=select_save_file,
        bg='#E5F7F5', fg=TEAL_DARK,
        activebackground='#D5F0ED', activeforeground=TEAL_DARK,
        font=('Microsoft YaHei UI', 8, 'bold'), relief='flat',
        bd=0, padx=14, pady=9, cursor='hand2', takefocus=True
    )
    output_button.grid(row=0, column=1, sticky='ns')

    # 固定像素高度的全宽主按钮，避免高 DPI 下文字裁切。
    action_host = tk.Frame(workbench, bg=SURFACE, height=48)
    action_host.pack(fill='x', pady=(0, 8))
    action_host.pack_propagate(False)
    gen_btn = tk.Button(
        action_host, text='生成并导出 Excel',
        command=lambda: on_generate(), bg=TEAL, fg=WHITE,
        activebackground=TEAL_DARK, activeforeground=WHITE,
        disabledforeground='#F1F6F5', font=('Microsoft YaHei UI', 10, 'bold'),
        relief='flat', bd=0, cursor='hand2', takefocus=True
    )
    gen_btn.place(x=0, y=0, relwidth=1, height=46)

    progress = ttk.Progressbar(
        workbench, mode='indeterminate',
        style='Reference.Horizontal.TProgressbar'
    )
    progress.pack(fill='x', pady=(0, 6))
    progress.pack_forget()

    log_panel = tk.Frame(
        workbench, bg=SURFACE_ALT, highlightthickness=1,
        highlightbackground=BORDER, bd=0
    )
    log_panel.pack(fill='both', expand=True)
    log_header = tk.Frame(log_panel, bg=SURFACE_ALT, height=30)
    log_header.pack(fill='x')
    log_header.pack_propagate(False)
    tk.Label(
        log_header, text='●  运行日志', bg=SURFACE_ALT, fg=TEAL_DARK,
        font=('Microsoft YaHei UI', 8, 'bold')
    ).pack(side='left', padx=10)
    tk.Label(
        log_header, textvariable=action_status_var, bg=SURFACE_ALT, fg=MUTED,
        font=('Microsoft YaHei UI', 8, 'bold')
    ).pack(side='right', padx=10)

    log_text = tk.Text(
        log_panel, font=('Cascadia Mono', 8, 'bold'), height=4, wrap='word',
        bg=SURFACE_ALT, fg='#446765', insertbackground=TEXT,
        selectbackground='#CBEAE7', relief='flat', bd=0,
        padx=10, pady=6, highlightthickness=0
    )
    log_text.pack(fill='both', expand=True)
    log_text.tag_configure('error', foreground='#C93645')
    log_text.tag_configure(
        'alert', foreground='#C62828', font=('Cascadia Mono', 8, 'bold')
    )
    log_text.tag_configure('success', foreground=SUCCESS)
    log_text.insert('end', '请选择两个 Excel 文件开始处理\n')
    log_text.configure(state='disabled')

    def clear_log():
        log_text.configure(state='normal')
        log_text.delete('1.0', 'end')
        log_text.configure(state='disabled')

    def append_log(msg):
        msg = str(msg)
        tag = ''
        if msg.startswith('[异常提醒]'):
            tag = 'alert'
        elif '错误' in msg or '失败' in msg:
            tag = 'error'
        elif '完成' in msg:
            tag = 'success'
        log_text.configure(state='normal')
        log_text.insert('end', msg + '\n', tag)
        log_text.see('end')
        log_text.configure(state='disabled')

    def log(msg):
        log_queue.put(msg)

    def poll_log_queue():
        try:
            while True:
                append_log(log_queue.get_nowait())
        except queue.Empty:
            pass
        root.after(80, poll_log_queue)

    interactive_widgets = [
        *tile_buttons, date_entry, output_entry, output_button, gen_btn
    ]

    def show_success(out, n, regions):
        output_var.set(out)
        status_var.set('生成完成')
        action_status_var.set(f'完成 · {n} 家门店 / {regions} 个区域')
        messagebox.showinfo(
            '生成完成',
            f'统计表已生成：\n{out}\n\n共 {n} 家门店，{regions} 个区域'
        )

    def show_failure(message):
        status_var.set('生成失败')
        action_status_var.set('请查看运行日志')
        messagebox.showerror('生成失败', message)

    def finish_generation():
        for widget in interactive_widgets:
            widget.configure(state='normal')
        gen_btn.configure(text='生成并导出 Excel', bg=TEAL)
        progress.stop()
        progress.pack_forget()
        generating['state'] = False

    def on_generate():
        if generating['state']:
            return
        dd_path = dd_var.get().strip()
        rl_path = rl_var.get().strip()
        out_path = output_var.get().strip() or None
        d_str = date_var.get().strip() or None

        if not dd_path or not os.path.exists(dd_path):
            messagebox.showerror('文件无效', '请选择有效的钉钉记录文件')
            return
        if not rl_path or not os.path.exists(rl_path):
            messagebox.showerror('文件无效', '请选择有效的日流水文件')
            return

        clear_log()
        for widget in interactive_widgets:
            widget.configure(state='disabled')
        gen_btn.configure(text='正在合并数据…', bg='#8FBAB6')
        status_var.set('正在处理')
        action_status_var.set('正在读取并自动匹配数据')
        progress.pack(fill='x', pady=(0, 6), before=log_panel)
        progress.start(12)
        generating['state'] = True

        def task():
            try:
                out, n, regions = run_pipeline(
                    dd_path, rl_path, out_path, d_str, log=log
                )
                if out:
                    root.after(
                        0, lambda out=out, n=n, regions=regions:
                        show_success(out, n, regions)
                    )
                else:
                    root.after(
                        0, lambda: show_failure('生成失败，请查看运行日志')
                    )
            except Exception as error:
                log(f'\n错误: {error}')
                root.after(0, lambda err=error: show_failure(str(err)))
            finally:
                root.after(0, finish_generation)

        threading.Thread(target=task, daemon=True).start()

    root.bind('<Control-Return>', lambda _event: on_generate())
    root.bind(
        '<Escape>',
        lambda _event: root.destroy() if not generating['state'] else None
    )
    root.after(80, poll_log_queue)
    root.mainloop()


if __name__ == '__main__':
    # 不带 --cli 参数时启动 GUI
    if '--cli' in sys.argv:
        cli_main()
    else:
        gui_main_reference()
